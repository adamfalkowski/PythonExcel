from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import * 
import os.path
import re




def transfer(operation, original_workbook_name, new_workbook, new_worksheet_name, spreadsheet_name):
    
    original_workbook = load_workbook(original_workbook_name)
    original_workbook_sheet = original_workbook.active

    column_headers = []

    new_workbook_sheet = new_workbook.active

    # --------------------------------------------------------------------------------
    #region:Definied Names
    # --------------------------------------------------------------------------------
    if operation == "defined names":

        orignal_def_names = original_workbook.defined_names.definedName #List of defined names

        for name in orignal_def_names:
            
            if name.value != "#REF!":
                _ , second_value = name.value.split("!")
                new_def_names = DefinedName(name=name.name,attr_text=f"{new_worksheet_name}!{second_value}")
                new_workbook.defined_names.append(new_def_names)
        
        new_workbook.save(spreadsheet_name)
    #endregion

    # --------------------------------------------------------------------------------
    #region: Defined names of Column Headers
    # --------------------------------------------------------------------------------
    elif operation == "defined names column headers":

        orignal_def_names = original_workbook.defined_names.definedName #List of defined names
        for name in orignal_def_names:
            
            if name.value != "#REF!":
                column_headers.append(name.name) 
                
        return column_headers
    #endregion

    # --------------------------------------------------------------------------------
    #region: Column Headers
    # --------------------------------------------------------------------------------
    elif operation == "column headers":

        for cell in original_workbook_sheet[1]: 
            new_workbook_sheet[cell.coordinate] = cell.value
            

        new_workbook.save(spreadsheet_name) 
        
    else:
        return 0
    #endregion

def collect(faults_wb, tags_equipment_wb):

    scada_ws = faults_wb["SCADA"]
    tags_equipment_ws = tags_equipment_wb["Tags_Equipment"]

    # --------------------------------------------------------------------------------
    #region:Initialization
    # --------------------------------------------------------------------------------
    default_value = ""
    column_names  = []
    scada_elements = []
    maint_units = []
    cameras = []
    num_maint_units = 0
    num_cameras = 0
    #endregion

    # --------------------------------------------------------------------------------
    #region:Create List of Columns from TagsEquipment:
    # --------------------------------------------------------------------------------
    for row in tags_equipment_ws.iter_rows(min_row = 1, max_row = 1):
        for header in row:
            column_names.append(header.value)
    #endregion
            
    # --------------------------------------------------------------------------------
    #region: Search for first row with non empty cell in column 4.  Ignore the first row beacuse it may contain a header:
    # --------------------------------------------------------------------------------
    first_row = 1
    for row in scada_ws.iter_rows(min_row = 2,values_only=True):
        first_row += 1
        if row[4] != None:
            break
    #endregion
        
    # --------------------------------------------------------------------------------
    #region:Default Values:
    # -------------------------------------------------------------------------------- 
    base_element = {"System": "BMS", 
                    "PLC/Link": "0_1", 
                    "OPC Prefix": "BSC",
                    "Alarm Filter Tree": "SCADAViewDetailed",
                    "ScanRate" : "1500",
                    "Edge Color": "Black", 
                    "Level 2 View" : "SCADAViewDetailed",
                    "LeftClick": "NoAction"}
    
    #endregion

    for row in scada_ws.iter_rows(min_row=first_row, values_only=True):
        if row[4] != None:

            # --------------------------------------------------------------------------------
            # Maintenance Unit
            # --------------------------------------------------------------------------------
            if re.search(r"^Maint.*Unit.*\d+$",row[4]):
                maint_unit = {key: default_value for key in column_names} # Create a dictionary for each Maint Unit
                maint_unit |= base_element # dictionary merge operation for default values
                num_maint_units += 1
                element_number = (re.findall(r'\d+',row[0]))[0] # return a list of matches, but only have one
                
                maint_unit["mcid"] = num_maint_units
                maint_unit["Equipment Line"] = f"MAINTENANCE_UNIT_Z0{num_maint_units}_UNIT" if num_maint_units < 10 else f"MAINTENANCE_UNIT_Z{num_maint_units}_UNIT"   
                maint_unit["Display Name"] = f"+Z0{num_maint_units}.2" if num_maint_units < 10 else f"+Z{num_maint_units}.2"
                maint_unit["Equipment Element"] = "Maintenance_unit"
                maint_unit["Layer"] = "MaintenanceUnit"
                maint_unit["TypeDescription"] = "Maintenance_unit"
                maint_unit["StyleIdentifier"] = "BM_MaintenanceUnit"
                maint_unit["Signal Mapping 1"] = "BM_MaintenanceUnit"
                maint_unit["Signal Address 1"] = f"DB_SCADA_FXG.tInterface.tStatus.awMisc_Element[{element_number}]"

            
                maint_units.append(maint_unit)
                
            # --------------------------------------------------------------------------------
            # Cameras
            # --------------------------------------------------------------------------------
            elif re.search(r"^Camera.*\d+$",row[4]):
                camera = {key: default_value for key in column_names} # Create a dictionary for each Maint Unit
                camera |= base_element
                num_cameras += 1
                element_number = (re.findall(r'\d+',row[0]))[0]
                camera["mcid"] = num_cameras
                camera["Equipment Line"] = f"CAMERA_{num_cameras}"  
                camera["Display Name"] = f"IN0{num_cameras}D10+10-VML"
                camera["Layer"] = "Camera"
                camera["TypeDescription"] = "Camera"
                camera["StyleIdentifier"] = "BG_ScannerGeneric"
                camera["Signal Mapping 1"] = "BM_UnitDefault"
                camera["Signal Address 1"] = f"DB_SCADA_FXG.tInterface.tStatus.awMisc_Element[{element_number}]"

                cameras.append(camera)


    scada_elements.extend([maint_units, cameras])

    return scada_elements

    
def send(tags_equipment_wb,scada_elements): 
    tags_equipment_ws = tags_equipment_wb["Tags_Equipment"]

    def get_column_index(key):
        for cells in tags_equipment_ws.iter_rows(min_row = 1, max_row = 1):
            for cell in cells:
                if cell.value == key:
                    return cell.column - 1 

    # Intitially clear all rows            
    for row in tags_equipment_ws.iter_rows(min_row = 2):
        for cell in row:
            cell.value = None           

    for elements in scada_elements:
        for element in elements:
            row = [''] * 16
            for key, value in element.items():
                row.insert(get_column_index(key), value)
            tags_equipment_ws.append(row)
    tags_equipment_wb.save("TagsEquipment_6200000120.xlsx") 


        
    


def main():
    folder_path = "D:\Practice\Python"

    #region: Load Spreadsheets:

    # --------------------------------------------------------------------------------
    # Load Tags Equipment
    # --------------------------------------------------------------------------------
    tags_equipment_wb_name = "TagsEquipment_6200000120.xlsx"
    tags_equipment_wb_path = folder_path + tags_equipment_wb_name
    
    # Create Tags Equipment if it does not exist
    if not os.path.exists(tags_equipment_wb_path):
        tags_equipment_wb = Workbook()
        tags_equipment_ws = tags_equipment_wb.active
        tags_equipment_ws.title = "Tags_Equipment"
        tags_equipment_wb.save(tags_equipment_wb_name) 
    else:
        tags_equipment_wb = load_workbook(tags_equipment_wb_path)

    # --------------------------------------------------------------------------------
    # Load Faults
    # --------------------------------------------------------------------------------
    file_path_faults = "D:\\617000101_FedEx_OLIV\sort9811\\bsc\Excel\Faults_9811.xlsm"  
    faults_wb = load_workbook(file_path_faults)
    #endregion

    transfer("defined names","TagsEquipment_bgfcv_templete.xlsx", tags_equipment_wb, "Tags_Equipment", tags_equipment_wb_name)
    transfer("defined names column headers","TagsEquipment_bgfcv_templete.xlsx",tags_equipment_wb, "Tags_Equipment", tags_equipment_wb_name)
    transfer("column headers","TagsEquipment_bgfcv_templete.xlsx",tags_equipment_wb, "Tags_Equipment", tags_equipment_wb_name)

    scada_elements = collect(faults_wb, tags_equipment_wb)
    send(tags_equipment_wb,scada_elements)

main()
    
